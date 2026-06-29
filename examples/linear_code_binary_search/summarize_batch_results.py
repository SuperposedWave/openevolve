#!/usr/bin/env python3
"""Summarize batch sweep results into Markdown and Excel."""

from __future__ import annotations

import argparse
import json
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_SUMMARY_DIR = SCRIPT_DIR / "outputs" / "_summaries" / "latest"
DEFAULT_SUMMARY = DEFAULT_SUMMARY_DIR / "summary.jsonl"
DEFAULT_MD = DEFAULT_SUMMARY_DIR / "experiment_summary.md"
DEFAULT_XLSX = DEFAULT_SUMMARY_DIR / "experiment_summary.xlsx"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export batch sweep results as Markdown and Excel.")
    parser.add_argument("--summary", default=str(DEFAULT_SUMMARY), help="Path to summary.jsonl")
    parser.add_argument("--markdown", default=str(DEFAULT_MD), help="Output Markdown path")
    parser.add_argument("--excel", default=str(DEFAULT_XLSX), help="Output Excel path")
    return parser.parse_args()


def load_rows(summary_path: Path) -> list[dict]:
    return [json.loads(line) for line in summary_path.read_text().splitlines() if line.strip()]


def build_markdown(rows: list[dict]) -> str:
    completed = [row for row in rows if row.get("status") == "completed"]
    successful = [row for row in completed if row.get("combined_score") == 1.0]
    partial = [row for row in completed if row.get("combined_score") not in (None, 1.0)]
    skipped = [row for row in rows if row.get("status") != "completed"]

    by_n: dict[int, dict[str, object]] = defaultdict(
        lambda: {"total": 0, "success": 0, "partial": 0, "partial_examples": []}
    )
    for row in completed:
        entry = by_n[row["n"]]
        entry["total"] += 1
        if row.get("combined_score") == 1.0:
            entry["success"] += 1
        else:
            entry["partial"] += 1
            if len(entry["partial_examples"]) < 4:
                entry["partial_examples"].append(
                    f"k={row['k']} (d={row['d']}, score={row['combined_score']:.4f})"
                )

    hardest_successes = sorted(successful, key=lambda row: (row["n"], row["d"], row["k"]))[-10:]
    hardest_partial = sorted(partial, key=lambda row: (row["n"], row["k"]))[-10:]

    lines = [
        "# Batch Sweep Summary",
        "",
        "## Experiment Setup",
        "",
        "- Search task: construct binary systematic parity-check matrices of the form `H = [P^T | I_r]`.",
        "- Priority source: OpenEvolve evolves a static `priority(column_mask, n, k, d)` heuristic.",
        "- Fixed search skeleton: enumerate legal candidate columns, score each column once, sort by score, then greedily accept every candidate that preserves the `d-1` column independence constraint.",
        "- Parameter source: `Misc/ECCRecord.json`.",
        "- Sweep window: `10 < n <= 20`.",
        "- Distance target per `(n, k)`: `lower` from `ECCRecord.json`.",
        "- Per-instance budget: `40` OpenEvolve iterations with event-based early stopping at `combined_score = 1.0`.",
        "- Validation: each run exports `matrix_verification.txt`; `complete` means the construction filled all `k` free columns, `partial` means the run stopped before filling them.",
        "",
        "## Overall Outcome",
        "",
        f"- Valid instances run: `{len(completed)}`",
        f"- Successful constructions at the requested lower bound: `{len(successful)}`",
        f"- Partial constructions: `{len(partial)}`",
        f"- Skipped invalid entries from the record (`k = n`): `{len(skipped)}`",
        f"- Success rate over valid instances: `{len(successful) / len(completed):.2%}`",
        "",
        "## By n",
        "",
        "| n | valid (k) | success | partial | representative partial cases |",
        "| --- | ---: | ---: | ---: | --- |",
    ]

    for n in sorted(by_n):
        entry = by_n[n]
        lines.append(
            f"| {n} | {entry['total']} | {entry['success']} | {entry['partial']} | "
            f"{', '.join(entry['partial_examples']) or '-'} |"
        )

    lines.extend(
        [
            "",
            "## Observed Pattern",
            "",
            "- For `n = 11` through `19`, the sweep succeeds for `k = 1` up to `k = n - 4`, and only the last three `k` values fail to reach the requested lower bound.",
            "- For `n = 20`, the sweep succeeds for `k = 1` through `15`, and fails for `k = 16, 17, 18, 19`.",
            "- Every partial case in this batch has target distance `d = 2`, so the misses concentrate in the very high-rate regime where `k` is close to `n`.",
            "- Partial verification still reports distance `2`, but the constructor did not manage to place all required free columns, so those instances do not count as successful lower-bound witnesses.",
            "",
            "## Largest Successful Instances",
            "",
            "| n | k | d | score | verification |",
            "| --- | ---: | ---: | ---: | --- |",
        ]
    )

    for row in hardest_successes:
        lines.append(
            f"| {row['n']} | {row['k']} | {row['d']} | {row['combined_score']:.4f} | {row.get('verification_status', '-')} |"
        )

    lines.extend(
        [
            "",
            "## Largest Partial Instances",
            "",
            "| n | k | d | score | constructed_columns | verification |",
            "| --- | ---: | ---: | ---: | ---: | --- |",
        ]
    )
    for row in hardest_partial:
        lines.append(
            f"| {row['n']} | {row['k']} | {row['d']} | {row['combined_score']:.4f} | "
            f"{int(row.get('constructed_columns', 0))} | {row.get('verification_status', '-')} |"
        )

    return "\n".join(lines) + "\n"


def write_excel(rows: list[dict], output_path: Path) -> None:
    completed = [row for row in rows if row.get("status") == "completed"]
    n_values = sorted({row["n"] for row in completed})
    k_values = sorted({row["k"] for row in completed})

    lookup = {(row["n"], row["k"]): row for row in completed}

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "lower_bound_results"

    sheet.cell(row=1, column=1, value="n \\ k")
    for col_idx, k in enumerate(k_values, start=2):
        sheet.cell(row=1, column=col_idx, value=k)

    success_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
    partial_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    invalid_fill = PatternFill(fill_type="solid", fgColor="E7E6E6")

    for row_idx, n in enumerate(n_values, start=2):
        sheet.cell(row=row_idx, column=1, value=n)
        for col_idx, k in enumerate(k_values, start=2):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if k >= n:
                cell.value = ""
                cell.fill = invalid_fill
                continue
            row = lookup.get((n, k))
            if row is None:
                cell.value = ""
                continue
            if row.get("combined_score") == 1.0:
                cell.value = "yes"
                cell.fill = success_fill
            else:
                cell.value = float(row["combined_score"])
                cell.fill = partial_fill
                cell.number_format = "0.0000"
            cell.alignment = Alignment(horizontal="center")

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    for row_idx in range(2, len(n_values) + 2):
        sheet.cell(row=row_idx, column=1).font = Font(bold=True)
        sheet.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center")

    sheet.freeze_panes = "B2"

    legend = workbook.create_sheet("legend")
    legend["A1"] = "Cell value"
    legend["B1"] = "Meaning"
    legend["A2"] = "yes"
    legend["B2"] = "Found a full construction meeting the requested lower bound"
    legend["A3"] = "0 < score < 1"
    legend["B3"] = "Did not fill all k free columns; score equals constructed_columns / k"
    legend["A4"] = "blank gray cell"
    legend["B4"] = "Invalid entry because k >= n was skipped"
    for cell in legend[1]:
        cell.font = Font(bold=True)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def main() -> None:
    args = parse_args()
    rows = load_rows(Path(args.summary).resolve())
    markdown_path = Path(args.markdown).resolve()
    excel_path = Path(args.excel).resolve()

    markdown_path.write_text(build_markdown(rows))
    write_excel(rows, excel_path)

    print(f"Markdown summary written to {markdown_path}")
    print(f"Excel summary written to {excel_path}")


if __name__ == "__main__":
    main()
